"""
Step 5: Export location statistics to Excel
- Reads from final_user_locations/{city}_final.json
- Outputs Excel file with each city as a separate sheet
- Sorted by total users (high to low)
"""

import json
import os
import sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============== Config ==============
INPUT_DIR = "final_user_locations"
OUTPUT_DIR = "final_user_locations"

# US state abbreviations
US_STATE_ABBREVS = {'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA',
                   'HI', 'ID', 'IL', 'IN', 'IA', 'KS', 'KY', 'LA', 'ME', 'MD',
                   'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ',
                   'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC',
                   'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY', 'DC'}

STATE_NAME_TO_ABBREV = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
    'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV',
    'new hampshire': 'NH', 'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
    'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK',
    'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
    'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
    'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
    'wisconsin': 'WI', 'wyoming': 'WY', 'district of columbia': 'DC'
}

# Invalid location words to filter out
INVALID_WORDS = {'earth', 'world', 'global', 'planet', 'everywhere', 'nowhere',
                 'internet', 'usa', 'america', 'refugee', 'heaven', 'hell', 'home'}

# ============== Utils ==============

def safe_print(text):
    """Safe print for Windows console"""
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode('ascii', 'replace').decode('ascii'))
    sys.stdout.flush()

def parse_location(loc):
    """Parse location string into (city_with_state, state, country)"""
    if not loc:
        return None, None, None

    parts = [p.strip() for p in loc.split(',')]
    city, state, country = None, None, None

    # Check if US location
    has_us_state_abbrev = any(p in US_STATE_ABBREVS for p in parts)
    has_us_state_name = any(p.lower() in STATE_NAME_TO_ABBREV for p in parts)
    has_usa = any(p in {'USA', 'United States', 'US'} for p in parts)

    if has_us_state_abbrev or has_us_state_name or has_usa:
        country = 'USA'
        for i, p in enumerate(parts):
            if p in US_STATE_ABBREVS:
                state = p
                if i > 0:
                    city = parts[0] + ', ' + state
                break
            elif p.lower() in STATE_NAME_TO_ABBREV:
                state = STATE_NAME_TO_ABBREV[p.lower()]
                if i > 0:
                    city = parts[0] + ', ' + state
                break
    else:
        if len(parts) >= 2:
            country = parts[-1]
            city = parts[0] + ', ' + country

    # Filter invalid cities
    if city and any(w in city.lower() for w in INVALID_WORDS):
        city = None

    return city, state, country

def load_json(filepath):
    """Load JSON file if exists"""
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

# ============== Main ==============

def export_excel():
    """Export Excel file with each city as a separate sheet"""
    import glob

    safe_print(f"\n{'='*60}")
    safe_print(f"EXPORT EXCEL")
    safe_print(f"{'='*60}")

    # Find all final.json files
    final_files = glob.glob(os.path.join(INPUT_DIR, "*_final.json"))

    if not final_files:
        safe_print(f"  No final.json files found in {INPUT_DIR}")
        return

    safe_print(f"  Found {len(final_files)} city files")

    star_sources = {'star-friend-analysis'}

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Process each city
    for final_file in sorted(final_files):
        filename = os.path.basename(final_file)
        source_city = filename.replace('_final.json', '')

        safe_print(f"  Processing {source_city}...")
        final_results = load_json(final_file)

        # Aggregate by city and state
        city_stats = {}
        state_stats = {}

        for user in final_results.values():
            loc = user.get('location')
            if not loc:
                continue

            source = user.get('source', '')
            followers = user.get('followers_count', 0)
            is_star = source in star_sources

            city, state, country = parse_location(loc)

            # Aggregate by city
            if city:
                if city not in city_stats:
                    city_stats[city] = {'regular_users': 0, 'regular_followers': 0,
                                        'star_users': 0, 'star_followers': 0}
                if is_star:
                    city_stats[city]['star_users'] += 1
                    city_stats[city]['star_followers'] += followers
                else:
                    city_stats[city]['regular_users'] += 1
                    city_stats[city]['regular_followers'] += followers

            # Aggregate by state
            if state:
                if state not in state_stats:
                    state_stats[state] = {'regular_users': 0, 'regular_followers': 0,
                                          'star_users': 0, 'star_followers': 0}
                if is_star:
                    state_stats[state]['star_users'] += 1
                    state_stats[state]['star_followers'] += followers
                else:
                    state_stats[state]['regular_users'] += 1
                    state_stats[state]['regular_followers'] += followers

        # Sort by total users
        sorted_cities = sorted(city_stats.items(),
                              key=lambda x: x[1]['regular_users'] + x[1]['star_users'],
                              reverse=True)
        sorted_states = sorted(state_stats.items(),
                              key=lambda x: x[1]['regular_users'] + x[1]['star_users'],
                              reverse=True)

        # Create sheet for this city
        sheet_name = source_city[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # Write States section
        ws.append(['US States'])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append(['State', 'Regular Users', 'Regular Followers', 'Star Users', 'Star Followers', 'Total Users', 'Total Followers'])
        for col in range(1, 8):
            ws.cell(row=2, column=col).font = header_font
            ws.cell(row=2, column=col).fill = header_fill

        for state, stats in sorted_states:
            total_users = stats['regular_users'] + stats['star_users']
            total_followers = stats['regular_followers'] + stats['star_followers']
            ws.append([state, stats['regular_users'], stats['regular_followers'],
                      stats['star_users'], stats['star_followers'], total_users, total_followers])

        # Add empty row
        states_end = len(sorted_states) + 3
        ws.append([])

        # Write Cities section
        cities_start = states_end + 1
        ws.cell(row=cities_start, column=1, value='Cities')
        ws.cell(row=cities_start, column=1).font = Font(bold=True, size=14)

        header_row = cities_start + 1
        headers = ['City', 'Regular Users', 'Regular Followers', 'Star Users', 'Star Followers', 'Total Users', 'Total Followers']
        for col, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col, value=header)
            ws.cell(row=header_row, column=col).font = header_font
            ws.cell(row=header_row, column=col).fill = header_fill

        for i, (city, stats) in enumerate(sorted_cities):
            row = header_row + 1 + i
            total_users = stats['regular_users'] + stats['star_users']
            total_followers = stats['regular_followers'] + stats['star_followers']
            ws.cell(row=row, column=1, value=city)
            ws.cell(row=row, column=2, value=stats['regular_users'])
            ws.cell(row=row, column=3, value=stats['regular_followers'])
            ws.cell(row=row, column=4, value=stats['star_users'])
            ws.cell(row=row, column=5, value=stats['star_followers'])
            ws.cell(row=row, column=6, value=total_users)
            ws.cell(row=row, column=7, value=total_followers)

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['A'].width = 30  # Wider for city names

        safe_print(f"    {len(sorted_states)} states, {len(sorted_cities)} cities")

    # Save workbook
    excel_file = os.path.join(OUTPUT_DIR, "location_stats.xlsx")
    wb.save(excel_file)
    safe_print(f"\n  Excel file: {excel_file}")

def main():
    safe_print("="*60)
    safe_print("STEP 5: EXPORT EXCEL")
    safe_print("="*60)

    export_excel()

    safe_print("\n" + "="*60)
    safe_print("STEP 5 DONE!")
    safe_print("="*60)

if __name__ == "__main__":
    main()
